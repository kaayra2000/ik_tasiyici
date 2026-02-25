import openpyxl

wb = openpyxl.load_workbook("DK_Tutanaklari_2026.xlsx")
ws = wb.active

print(f"Sheet Name: {ws.title}")
print(f"Ad Soyad (B3): {ws['B3'].value}")
print(f"TCKN (C3): {ws['C3'].value}")
print(f"Birim (D3): {ws['D3'].value}")
print(f"Ünvan (E3): {ws['E3'].value}")
print(f"Kademe (F3): {ws['F3'].value}")
print(f"Toplam Prim Günü K10: {ws['K10'].value}")
print(f"Z2 (Hizmet Grubu): {ws['Z2'].value}")
