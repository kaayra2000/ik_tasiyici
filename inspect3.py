import openpyxl

wb = openpyxl.load_workbook("docs/EN YENİ DK format1 v2.xlsx")
ws = wb.active

for r in range(1, 25):
    for c in range(1, 15):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            print(f"{openpyxl.utils.get_column_letter(c)}{r}: {val}")
