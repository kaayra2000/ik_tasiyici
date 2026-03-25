"""
Çıktı Excel dosyasını oluşturan modül.

Her personel için ``cikti_ornegi.xlsx`` şablonunu kopyalarak bir sayfa oluşturur
ve tüm sayfaları tek bir ``DK_Tutanaklari_2026.xlsx`` dosyasında birleştirir.

Sayfa doldurma mantığı **Strategy Pattern** ile soyutlanmıştır.
Farklı şablon versiyonları ``ExcelWriteStrategy`` arayüzünü implemente eden
sınıflar olarak tanımlanır ve ``ExcelWriterFactory`` aracılığıyla yaratılır.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List
from copy import copy

import openpyxl
import openpyxl.worksheet.worksheet
from openpyxl import Workbook

from src.config.constants import (
    DEFAULT_VERSION,
    MAX_SHEET_NAME_LEN,
    OUTPUT_FILENAME,
    TEMPLATE_PATH,
)
from src.core.excel_reader import Personel
from src.core.excel_write_strategy import ExcelWriteStrategy
from src.core.excel_writer_factory import ExcelWriterFactory

# ---------------------------------------------------------------------------
# Ana yazma fonksiyonları
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class TutanakOlusturmaRaporu:
    """Tutanak oluşturma işleminin özetini taşır."""

    output_path: Path
    added_file_count: int = 0
    skipped_existing_file_count: int = 0
    generated_files: list[Path] = field(default_factory=list)
    warning_messages: list[str] = field(default_factory=list)

    @property
    def added_sheet_count(self) -> int:
        """Geriye dönük uyumluluk için dosya sayısını sayfa sayısı gibi sunar."""
        return self.added_file_count

    @property
    def skipped_existing_count(self) -> int:
        """Geriye dönük uyumluluk için atlanan dosya sayısını sunar."""
        return self.skipped_existing_file_count


@dataclass(frozen=True)
class _PersonelDosyaYazimSonucu:
    """Tek personel için dosyaya yazım sonucunu taşır."""

    output_path: Path
    dosyaya_yazildi: bool
    skipped_existing_count: int = 0
    warning_messages: list[str] = field(default_factory=list)


def olustur_dk_klasoru_raporlu(
    personeller: List[Personel],
    cikti_klasoru: str | Path,
    template_path: str | Path | None = None,
    version: str = DEFAULT_VERSION,
) -> TutanakOlusturmaRaporu:
    """Her personel için ayrı bir tutanak dosyası üretir.

    Dosya adı üretiminde sayfa adı kuralları kullanılır; yani
    ``{Ad Soyad} - {TCKN}.xlsx`` formatı uygulanır.
    """
    strategy = ExcelWriterFactory.create(version)
    cikti_klasoru = _hazirla_cikti_klasoru(cikti_klasoru)

    added_file_count = 0
    skipped_existing_file_count = 0
    warning_messages: list[str] = []
    generated_files: list[Path] = []

    for personel in personeller:
        personel_sonucu = _personel_dosyasina_yaz(
            personel=personel,
            cikti_klasoru=cikti_klasoru,
            strategy=strategy,
            template_path=template_path,
        )
        if personel_sonucu.dosyaya_yazildi:
            added_file_count += 1
            generated_files.append(personel_sonucu.output_path)
        skipped_existing_file_count += personel_sonucu.skipped_existing_count
        warning_messages.extend(personel_sonucu.warning_messages)

    return TutanakOlusturmaRaporu(
        output_path=cikti_klasoru,
        added_file_count=added_file_count,
        skipped_existing_file_count=skipped_existing_file_count,
        generated_files=generated_files,
        warning_messages=warning_messages,
    )


def _hazirla_cikti_klasoru(cikti_klasoru: str | Path) -> Path:
    """Klasörü Path'e çevirir ve yoksa oluşturur."""
    klasor = Path(cikti_klasoru)
    klasor.mkdir(parents=True, exist_ok=True)
    return klasor


def _personel_dosyasina_yaz(
    *,
    personel: Personel,
    cikti_klasoru: Path,
    strategy: ExcelWriteStrategy,
    template_path: str | Path | None,
) -> _PersonelDosyaYazimSonucu:
    """Tek personel için hedef dosyaya yazım akışını yürütür."""
    cikti_yolu = _personel_cikti_dosya_yolu(personel, cikti_klasoru)

    if cikti_yolu.exists():
        return _mevcut_dosyaya_personel_yaz(
            cikti_yolu=cikti_yolu,
            personel=personel,
            strategy=strategy,
            template_path=template_path,
        )

    return _yeni_personel_dosyasi_olustur(
        cikti_yolu=cikti_yolu,
        personel=personel,
        strategy=strategy,
        template_path=template_path,
    )


def _personel_cikti_dosya_yolu(personel: Personel, cikti_klasoru: Path) -> Path:
    """Personel için çıktı dosyasının tam yolunu üretir."""
    base_name = _sayfa_adi_olustur(personel)
    return cikti_klasoru / f"{base_name}.xlsx"


def _mevcut_dosyaya_personel_yaz(
    *,
    cikti_yolu: Path,
    personel: Personel,
    strategy: ExcelWriteStrategy,
    template_path: str | Path | None,
) -> _PersonelDosyaYazimSonucu:
    """Mevcut workbook'a personel sayfası ekler."""
    wb = openpyxl.load_workbook(cikti_yolu)
    added_count, skipped_count, warning_messages = _personelleri_workbooka_ekle(
        wb,
        [personel],
        strategy,
        template_path,
    )
    wb.save(cikti_yolu)
    wb.close()

    return _PersonelDosyaYazimSonucu(
        output_path=cikti_yolu,
        dosyaya_yazildi=bool(added_count),
        skipped_existing_count=skipped_count,
        warning_messages=warning_messages,
    )


def _yeni_personel_dosyasi_olustur(
    *,
    cikti_yolu: Path,
    personel: Personel,
    strategy: ExcelWriteStrategy,
    template_path: str | Path | None,
) -> _PersonelDosyaYazimSonucu:
    """Yeni workbook oluşturur ve personel sayfasını kaydeder."""
    wb, added_count, skipped_count, warning_messages = _workbook_olustur(
        [personel],
        strategy,
        template_path,
    )
    wb.save(cikti_yolu)
    wb.close()

    return _PersonelDosyaYazimSonucu(
        output_path=cikti_yolu,
        dosyaya_yazildi=bool(added_count),
        skipped_existing_count=skipped_count,
        warning_messages=warning_messages,
    )


def olustur_dk_dosyasi_raporlu(
    personeller: List[Personel],
    cikti_dizini: str | Path = ".",
    dosya_adi: str = OUTPUT_FILENAME,
    template_path: str | Path | None = None,
    version: str = DEFAULT_VERSION,
) -> TutanakOlusturmaRaporu:
    """
    Personel listesinden DK Tutanağı üretip ayrıntılı işlem raporu döner.

    :param personeller: İşlenecek personel listesi.
    :param cikti_dizini: Çıktı dosyasının kaydedileceği dizin.
    :param dosya_adi: Çıktı dosya adı.
    :param template_path: Özel çıktı şablonu yolu (opsiyonel).
    :param version: Çıktı versiyonu (ör. ``"v1"``).
    :returns: Oluşturma özeti.
    """
    strategy = ExcelWriterFactory.create(version)
    cikti_dizini = Path(cikti_dizini)
    cikti_dizini.mkdir(parents=True, exist_ok=True)
    cikti_yolu = cikti_dizini / dosya_adi

    if cikti_yolu.exists():
        wb = openpyxl.load_workbook(cikti_yolu)
        added_sheet_count, skipped_existing_count, warning_messages = (
            _personelleri_workbooka_ekle(wb, personeller, strategy, template_path)
        )
    else:
        wb, added_sheet_count, skipped_existing_count, warning_messages = (
            _workbook_olustur(personeller, strategy, template_path)
        )

    wb.save(cikti_yolu)
    wb.close()
    return TutanakOlusturmaRaporu(
        output_path=cikti_yolu,
        added_file_count=added_sheet_count,
        skipped_existing_file_count=skipped_existing_count,
        generated_files=[cikti_yolu],
        warning_messages=warning_messages,
    )


# ---------------------------------------------------------------------------
# İç yardımcılar
# ---------------------------------------------------------------------------


def _workbook_olustur(
    personeller: List[Personel],
    strategy: ExcelWriteStrategy,
    template_path: str | Path | None = None,
) -> tuple[Workbook, int, int, list[str]]:
    """Her personel için şablon sayfasından kopyalanmış Workbook oluşturur."""
    wb = Workbook()
    added_sheet_count, skipped_existing_count, warning_messages = (
        _personelleri_workbooka_ekle(wb, personeller, strategy, template_path)
    )

    # openpyxl en az 1 sayfa olmadan kaydedemez;
    # Eğer hiç personel sayfası oluşturulmadıysa varsayılan aktif sayfayı
    # silmek yerine `_bos` olarak yeniden adlandırıyoruz. Bu koruma
    # openpyxl'in workbook seviyesindeki varsayılan stillerinin korunmasını
    # sağlar (uyarıyı ortadan kaldırır) ve önceki davranışı taklit eder.
    if not wb.sheetnames:
        wb.create_sheet("_bos")
    elif (
        added_sheet_count == 0
        and len(wb.sheetnames) == 1
        and wb.sheetnames[0] == "Sheet"
    ):
        wb.active.title = "_bos"

    return wb, added_sheet_count, skipped_existing_count, warning_messages


def _personelleri_workbooka_ekle(
    wb: Workbook,
    personeller: List[Personel],
    strategy: ExcelWriteStrategy,
    template_path: str | Path | None = None,
) -> tuple[int, int, list[str]]:
    """Workbook'a yalnızca eksik personel sayfalarını sona ekler."""
    template_wb = None
    template_ws = None
    added_sheet_count = 0
    skipped_existing_count = 0
    warning_messages: list[str] = []

    try:
        for personel in personeller:
            sayfa_adi = _sayfa_adi_olustur(personel)
            if sayfa_adi in wb.sheetnames:
                skipped_existing_count += 1
                warning_messages.append(_build_skip_message(personel, sayfa_adi))
                continue

            if template_ws is None:
                template_wb = openpyxl.load_workbook(
                    _template_yolunu_coz(template_path)
                )
                if not template_wb.sheetnames:
                    raise ValueError("Şablon workbook içinde hiç sayfa yok.")
                template_ws = template_wb.active

            # Reuse the default active sheet for the first generated page to
            # preserve workbook-level defaults (openpyxl's default style).
            if (
                len(wb.sheetnames) == 1
                and wb.active is not None
                and not getattr(wb.active, "_cells", {})
            ):
                ws = wb.active
                ws.title = sayfa_adi
            else:
                ws = wb.create_sheet(title=sayfa_adi)
            _sayfa_icerigini_kopyala(template_ws, ws)
            strategy.sayfa_doldur(ws, personel)
            added_sheet_count += 1
    finally:
        if template_wb is not None:
            template_wb.close()

    return added_sheet_count, skipped_existing_count, warning_messages


def _template_yolunu_coz(template_path: str | Path | None = None) -> Path:
    """Şablon dosya yolunu çözümler ve dosyanın varlığını doğrular."""
    if template_path is None:
        project_root = Path(__file__).resolve().parent.parent.parent
        t_path = project_root / TEMPLATE_PATH
    else:
        t_path = Path(template_path)

    if not t_path.exists():
        raise FileNotFoundError(f"Excel şablonu bulunamadı: {t_path}")

    return t_path


def _boyut_ozelliklerini_kopyala(kaynak_boyut: object, hedef_boyut: object) -> None:
    """Boyutla alakalı özellikleri kopyalar"""
    alanlar = (
        "width",
        "hidden",
        "bestFit",
        "outlineLevel",
        "outline_level",
        "collapsed",
        "customWidth",
        "auto_size",
        "style",
        "min",
        "max",
        "height",
        "customHeight",
    )
    for alan in alanlar:
        if hasattr(kaynak_boyut, alan):
            try:
                setattr(hedef_boyut, alan, getattr(kaynak_boyut, alan))
            except AttributeError:
                # Bazı alanlar read-only olabilir (örn. customHeight). Güvenle atla.
                continue


def _kopyala_hucreler(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Hücre değerlerini, stillerini ve ek açıklamalarını kopyalar."""
    for (satir, sutun), kaynak_hucre in kaynak_ws._cells.items():
        hedef_hucre = hedef_ws.cell(row=satir, column=sutun)
        hedef_hucre._value = kaynak_hucre._value
        hedef_hucre.data_type = kaynak_hucre.data_type

        if kaynak_hucre.has_style:
            hedef_hucre.font = copy(kaynak_hucre.font)
            hedef_hucre.fill = copy(kaynak_hucre.fill)
            hedef_hucre.border = copy(kaynak_hucre.border)
            hedef_hucre.alignment = copy(kaynak_hucre.alignment)
            hedef_hucre.number_format = kaynak_hucre.number_format
            hedef_hucre.protection = copy(kaynak_hucre.protection)

        if kaynak_hucre.hyperlink:
            hedef_hucre._hyperlink = copy(kaynak_hucre.hyperlink)

        if kaynak_hucre.comment:
            hedef_hucre.comment = copy(kaynak_hucre.comment)


def _kopyala_satir_boyutlari(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Satır boyutlarını kopyalar."""
    for anahtar, boyut in kaynak_ws.row_dimensions.items():
        hedef_boyut = hedef_ws.row_dimensions[anahtar]
        _boyut_ozelliklerini_kopyala(boyut, hedef_boyut)


def _kopyala_sutun_boyutlari(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sütun boyutlarını kopyalar."""
    for anahtar, boyut in kaynak_ws.column_dimensions.items():
        hedef_boyut = hedef_ws.column_dimensions[anahtar]
        _boyut_ozelliklerini_kopyala(boyut, hedef_boyut)


def _kopyala_sayfa_ozellikleri(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sayfa düzeyi ayarlarını küçük sorumluluklara bölerek kopyalar."""
    _kopyala_sayfa_metin_ve_gorunum_ozellikleri(kaynak_ws, hedef_ws)
    _kopyala_birlesik_hucre_araliklari(kaynak_ws, hedef_ws)
    _kopyala_yazdirma_ve_koruma_ayarlari(kaynak_ws, hedef_ws)
    _kopyala_veri_dogrulamalari(kaynak_ws, hedef_ws)
    _kopyala_dondurulmus_bolme(kaynak_ws, hedef_ws)
    _kopyala_yazdirma_alani(kaynak_ws, hedef_ws)


def _kopyala_sayfa_metin_ve_gorunum_ozellikleri(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Sayfanın biçim, özellik ve görünüm verilerini kopyalar."""
    hedef_ws.sheet_format = copy(kaynak_ws.sheet_format)
    hedef_ws.sheet_properties = copy(kaynak_ws.sheet_properties)
    hedef_ws.views = copy(kaynak_ws.views)


def _kopyala_birlesik_hucre_araliklari(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Birleşik hücre aralıklarını güvenli biçimde yeniden oluşturur."""
    # Birleşik hücre aralıklarını doğrudan nesne kopyasıyla taşımak,
    # bazı Excel istemcilerinde kenarlık/hizalama bozulmalarına yol açabilir.
    # Aralıkları tek tek merge ederek güvenli şekilde yeniden oluşturuyoruz.
    for birlesik_aralik in list(hedef_ws.merged_cells.ranges):
        hedef_ws.unmerge_cells(str(birlesik_aralik))
    for birlesik_aralik in kaynak_ws.merged_cells.ranges:
        hedef_ws.merge_cells(str(birlesik_aralik))


def _kopyala_yazdirma_ve_koruma_ayarlari(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Yazdırma, sayfa düzeni ve koruma ayarlarını kopyalar."""
    hedef_ws.page_margins = copy(kaynak_ws.page_margins)
    hedef_ws.page_setup = copy(kaynak_ws.page_setup)
    hedef_ws.print_options = copy(kaynak_ws.print_options)
    hedef_ws.protection = copy(kaynak_ws.protection)
    hedef_ws.conditional_formatting = copy(kaynak_ws.conditional_formatting)


def _kopyala_veri_dogrulamalari(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Varsa veri doğrulama tanımlarını kopyalar."""
    if hasattr(kaynak_ws, "data_validations"):
        hedef_ws.data_validations = copy(kaynak_ws.data_validations)


def _kopyala_dondurulmus_bolme(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Freeze panes ayarını kopyalar."""
    hedef_ws.freeze_panes = kaynak_ws.freeze_panes


def _kopyala_yazdirma_alani(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Yazdırma alanını güvenli biçimde kopyalar."""
    # NOT: copy(PrintArea) nesnesi MultiCellRange'e dönüşür ve sayfa
    # nitelendirmesini kaybeder -> Excel bozuk Named Range hatası verir.
    # Public setter kullanarak doğru formatta yazdırma alanı oluşturulur.
    if kaynak_ws._print_area:
        hedef_ws.print_area = str(kaynak_ws._print_area)


def _sayfa_icerigini_kopyala(kaynak_ws: openpyxl.worksheet.worksheet.Worksheet, hedef_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Harici workbook'taki şablon sayfasını hedef workbook'a klonlar."""
    _kopyala_hucreler(kaynak_ws, hedef_ws)
    _kopyala_satir_boyutlari(kaynak_ws, hedef_ws)
    _kopyala_sutun_boyutlari(kaynak_ws, hedef_ws)
    _kopyala_sayfa_ozellikleri(kaynak_ws, hedef_ws)


def _sayfa_adi_olustur(personel: Personel) -> str:
    """
    Personel için Excel sayfa adını oluşturur.

    Format: ``{Ad Soyad} - {TCKN}`` (max 31 karakter).

    :param personel: Personel nesnesi.
    :returns: Excel sayfa adı.
    """
    tam_ad = f"{personel.ad_soyad} - {personel.tckn}"
    if len(tam_ad) > MAX_SHEET_NAME_LEN:
        kisaltilmis = personel.ad_soyad[: MAX_SHEET_NAME_LEN - len(personel.tckn) - 5]
        tam_ad = f"{kisaltilmis}.. - {personel.tckn}"
    for karakter in r"\/?*[]":
        tam_ad = tam_ad.replace(karakter, "")
    return tam_ad[:MAX_SHEET_NAME_LEN]


def _build_skip_message(personel: Personel, sayfa_adi: str) -> str:
    """Workbook'ta zaten bulunan kayıt için log mesajı üretir."""
    return (
        "Kayıt atlandı: hedef dosyada zaten mevcut. "
        f"SAYFA='{sayfa_adi}', "
        f"TCKN='{personel.tckn}', "
        f"AD SOYAD='{personel.ad_soyad}', "
        f"BİRİMİ='{personel.birim}'"
    )


