"""
V1 çıktı şablonu için Excel yazma stratejisi.

``cikti_taslagi.xlsx`` şablonuna uygun hücre adresleri, formül yazma
mantığı ve veri doğrulama kurallarını içerir.
Mevcut ``excel_writer.py`` içindeki iç fonksiyonlardan birebir taşınmıştır.
"""

from __future__ import annotations

from copy import copy

import openpyxl
import openpyxl.worksheet.worksheet

from src.config.constants import (
    TECRUBE_BASLANGIC_SATIR,
    TECRUBE_BITIS_SATIR,
)
from src.core.excel_reader import Personel
from src.core.excel_write_strategy import ExcelWriteStrategy
from src.core.formula_builder import (
    BRUT_UCRET_HARITASI,
    alanda_prim_formulu,
    brut_ucret_formulu,
    en_yuksek_ogrenim_formulu,
    hizmet_grubu_formulu,
    kademe_formulu,
    kademe_baslangic_formulu,
    kademe_bitis_formulu,
    prim_gunu_formulu,
    tecrube_360_ay_formulu,
    tecrube_360_gun_formulu,
    tecrube_360_yil_formulu,
    tecrube_yili_formulu,
    toplam_alanda_prim_formulu,
    toplam_prim_formulu,
    unvan_formulu,
)

# ---------------------------------------------------------------------------
# V1 şablonuna ait sabitler – hücre adresleri
# ---------------------------------------------------------------------------

# Otomatik doldurulacak (o) hücreler
_HUCRE_AD_SOYAD = "B3"
_HUCRE_TCKN = "C3"
_HUCRE_BIRIM = "D3"
_HUCRE_UNVAN = "E3"
_HUCRE_KADEME = "F3"
_HUCRE_BRUT_UCRET = "G3"
_UCRET_TABLO_ANAHTAR_SUTUNU = "AA"
_UCRET_TABLO_DEGER_SUTUNU = "AB"
_UCRET_TABLO_BASLANGIC_SATIR = 1
_HUCRE_HIZMET_GRUBU_BASLIK = "M2"
_HUCRE_HIZMET_GRUBU_TURU = "M3"
_HUCRE_HIZMET_GRUBU_BASLIK_STIL_KAYNAGI = "H2"
_HUCRE_HIZMET_GRUBU_TURU_STIL_KAYNAGI = "H3"
_HIZMET_GRUBU_SUTUNU = "M"
_HIZMET_GRUBU_MIN_SUTUN_GENISLIGI = 22
_HUCRE_EKSIK_GUN_BASLIK = "M12"
_HUCRE_EKSIK_GUN_BASLIK_STIL_KAYNAGI = "L12"
_HUCRE_EKSIK_GUN_STIL_KAYNAGI = "L13"
_HUCRE_EKSIK_GUN_BASLIK_METNI = "Eksik Gün Sayısı"
_EKSIK_GUN_SUTUNU = "M"

# Toplam / hesap satırları
_SATIR_TOPLAM_PRIM = TECRUBE_BITIS_SATIR + 1  # 28
_SATIR_ALANDA_PRIM = TECRUBE_BITIS_SATIR + 1  # 28
_SATIR_YIL_AY_GUN = TECRUBE_BITIS_SATIR + 2  # 29

# M3 görünür seçim hücresidir:
# M3 = Hizmet Grubu Türü (A / AG)
#
# Z sütununu gizli hesaplama için kullanıyoruz:
# Z1 = Tecrübe Yılı, Z2 = Hizmet Grubu, Z3 = Kademe, Z4 = En Yüksek Öğrenim
_TECRUBE_YILI_HUCRE = "Z1"
_EN_YUKSEK_OGRENIM_HUCRE = "Z4"

# Şablondaki öğrenim satırları parametreleri (B=Ad, C=Okul, K=Alanında)
_OGRENIM_BAS_SATIR = 6
_OGRENIM_BIT_SATIR = 10
_OGRENIM_AD_SUTUN = "B"
_OGRENIM_OKUL_SUTUN = "C"
_OGRENIM_ALANINDA_SUTUN = "K"


class ExcelWriteStrategyV1(ExcelWriteStrategy):
    """V1 (Standart DK Taslağı) için Excel yazma stratejisi.

    ``cikti_taslagi.xlsx`` / ``cikti_taslagi_dolu.xlsx`` şablonlarına
    uygun olarak sayfayı doldurur.
    """

    # V1 sabitlerini dışarıya da açıyoruz (test vb. için)
    HUCRE_UNVAN = _HUCRE_UNVAN

    def sayfa_doldur(self, ws: openpyxl.worksheet.worksheet.Worksheet, personel: Personel) -> None:
        """V1 şablonuna göre çalışma sayfasını doldurur."""
        self._doldur_otomatik(ws, personel)
        self._ekle_veri_dogrulama(ws)
        self._yaz_tecrube_satirlari(ws)
        self._yaz_hesap_satirlari(ws)

    # ------------------------------------------------------------------
    # İç yardımcılar
    # ------------------------------------------------------------------

    @staticmethod
    def _doldur_otomatik(ws: openpyxl.worksheet.worksheet.Worksheet, personel: Personel) -> None:
        """Otomatik (o) alanları personel verisinden doldurur."""
        ws[_HUCRE_AD_SOYAD] = personel.ad_soyad
        ws[_HUCRE_TCKN] = personel.tckn
        ws[_HUCRE_BIRIM] = personel.birim
        ws[_HUCRE_HIZMET_GRUBU_BASLIK] = "HİZMET GRUBU TÜRÜ"
        ws[_HUCRE_HIZMET_GRUBU_TURU] = "AG"
        ExcelWriteStrategyV1._kopyala_hucre_bicimi(
            ws,
            _HUCRE_HIZMET_GRUBU_BASLIK_STIL_KAYNAGI,
            _HUCRE_HIZMET_GRUBU_BASLIK,
        )
        ExcelWriteStrategyV1._kopyala_hucre_bicimi(
            ws,
            _HUCRE_HIZMET_GRUBU_TURU_STIL_KAYNAGI,
            _HUCRE_HIZMET_GRUBU_TURU,
        )
        ExcelWriteStrategyV1._ayarla_hizmet_grubu_sutun_genisligi(ws)

    @staticmethod
    def _yaz_tecrube_satirlari(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Her mesleki tecrübe satırı için Excel formüllerini yazar."""
        ws[_HUCRE_EKSIK_GUN_BASLIK] = _HUCRE_EKSIK_GUN_BASLIK_METNI
        ExcelWriteStrategyV1._kopyala_hucre_bicimi(
            ws,
            _HUCRE_EKSIK_GUN_BASLIK_STIL_KAYNAGI,
            _HUCRE_EKSIK_GUN_BASLIK,
        )
        ExcelWriteStrategyV1._ayarla_eksik_gun_sutun_genisligi(ws)

        for satir in range(TECRUBE_BASLANGIC_SATIR, TECRUBE_BITIS_SATIR + 1):
            ws.cell(row=satir, column=11).value = prim_gunu_formulu(satir)
            ws.cell(row=satir, column=12).value = alanda_prim_formulu(satir)
            ExcelWriteStrategyV1._kopyala_hucre_bicimi(
                ws,
                _HUCRE_EKSIK_GUN_STIL_KAYNAGI,
                f"{_EKSIK_GUN_SUTUNU}{satir}",
            )

    @staticmethod
    def _yaz_hesap_satirlari(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Toplam, tecrübe yılı, ünvan, hizmet grubu ve kademe satırlarını yazar."""
        toplam_alanda_hucre = f"L{_SATIR_ALANDA_PRIM}"

        # Toplam Prim Günü (K19) ve Alanda Toplam Prim Günü (L19)
        ws.cell(row=_SATIR_TOPLAM_PRIM, column=11).value = toplam_prim_formulu()
        ws.cell(row=_SATIR_ALANDA_PRIM, column=12).value = toplam_alanda_prim_formulu()

        ExcelWriteStrategyV1._yaz_360_yil_ay_gun(ws)

        # Z sütununa gizli formülleri yazalım:
        # Z1 = Tecrübe Yılı (alanda toplam prim / 360)
        ws["Z1"] = tecrube_yili_formulu(toplam_alanda_hucre)

        # Z4 = En yüksek alanında öğrenim
        ws[_EN_YUKSEK_OGRENIM_HUCRE] = en_yuksek_ogrenim_formulu(
            baslangic_satir=_OGRENIM_BAS_SATIR,
            bitis_satir=_OGRENIM_BIT_SATIR,
            ad_sutun=_OGRENIM_AD_SUTUN,
            okul_sutun=_OGRENIM_OKUL_SUTUN,
            alaninda_sutun=_OGRENIM_ALANINDA_SUTUN,
        )

        # Z2 = Hizmet Grubu
        ws["Z2"] = hizmet_grubu_formulu(
            _TECRUBE_YILI_HUCRE,
            _HUCRE_HIZMET_GRUBU_TURU,
        )

        # Z3 = Kademe
        ws["Z3"] = kademe_formulu(_TECRUBE_YILI_HUCRE, _EN_YUKSEK_OGRENIM_HUCRE)

        # E3: Ünvan
        ws[_HUCRE_UNVAN] = unvan_formulu(
            _TECRUBE_YILI_HUCRE,
            _HUCRE_HIZMET_GRUBU_TURU,
        )

        # F3: Derece/Kademe
        ws[_HUCRE_KADEME] = '=IF(Z3="", Z2, Z2 & "/" & Z3)'

        ExcelWriteStrategyV1._yaz_brut_ucret_tablosu(ws)

        # G3: Brüt Ücret (F3 değeri -> ücret tablosu)
        ilk_satir = _UCRET_TABLO_BASLANGIC_SATIR
        son_satir = _UCRET_TABLO_BASLANGIC_SATIR + len(BRUT_UCRET_HARITASI) - 1
        tablo_araligi = (
            f"${_UCRET_TABLO_ANAHTAR_SUTUNU}${ilk_satir}:"
            f"${_UCRET_TABLO_DEGER_SUTUNU}${son_satir}"
        )
        ws[_HUCRE_BRUT_UCRET] = brut_ucret_formulu(_HUCRE_KADEME, tablo_araligi)

        # K30: Kademe Başlangıcı
        k30_hucre = ws["K30"]
        k30_hucre.value = kademe_baslangic_formulu(
            _TECRUBE_YILI_HUCRE, _EN_YUKSEK_OGRENIM_HUCRE
        )

        # L30: Kademe Bitişi
        l30_hucre = ws["L30"]
        l30_hucre.value = kademe_bitis_formulu(
            _TECRUBE_YILI_HUCRE, _EN_YUKSEK_OGRENIM_HUCRE
        )

        # K30 ve L30'a tam sayı formatı uygula
        ExcelWriteStrategyV1._uygula_tam_sayi_formati(k30_hucre, l30_hucre)

    @staticmethod
    def _yaz_360_yil_ay_gun(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """J29/K29/L29: 30/360 bazlı toplam yıl/ay/gün formüllerini yazar."""
        yil_hucre = ws.cell(row=_SATIR_YIL_AY_GUN, column=10)
        ay_hucre = ws.cell(row=_SATIR_YIL_AY_GUN, column=11)
        gun_hucre = ws.cell(row=_SATIR_YIL_AY_GUN, column=12)
        yil_hucre.value = tecrube_360_yil_formulu()
        ay_hucre.value = tecrube_360_ay_formulu()
        gun_hucre.value = tecrube_360_gun_formulu()
        ExcelWriteStrategyV1._uygula_tam_sayi_formati(yil_hucre, ay_hucre, gun_hucre)

    @staticmethod
    def _uygula_tam_sayi_formati(*hucreler: openpyxl.cell.cell.Cell) -> None:
        """Verilen hücrelere tam sayı formatı (0) uygular.

        Virgülden sonraki kısımları göstermez (1.00 yerine 1 gösterir).
        """
        for hucre in hucreler:
            hucre.number_format = "0"

    @staticmethod
    def _kopyala_hucre_bicimi(ws: openpyxl.worksheet.worksheet.Worksheet, kaynak_hucre: str, hedef_hucre: str) -> None:
        """Hedef hücreye kaynak hücrenin biçimini uygular."""
        ws[hedef_hucre]._style = copy(ws[kaynak_hucre]._style)

    @staticmethod
    def _ayarla_hizmet_grubu_sutun_genisligi(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Başlık metninin görünmesi için hizmet grubu sütununu genişletir."""
        sutun = ws.column_dimensions[_HIZMET_GRUBU_SUTUNU]
        mevcut = sutun.width or 0
        sutun.width = max(mevcut, _HIZMET_GRUBU_MIN_SUTUN_GENISLIGI)

    @staticmethod
    def _ayarla_eksik_gun_sutun_genisligi(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Eksik gün başlığının görünmesi için M sütununu genişletir."""
        kaynak = ws.column_dimensions["L"]
        hedef = ws.column_dimensions[_EKSIK_GUN_SUTUNU]
        if kaynak.width and (hedef.width is None or hedef.width < kaynak.width):
            hedef.width = kaynak.width

    @staticmethod
    def _yaz_brut_ucret_tablosu(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Brut ucret esleme tablosunu gizli yardimci sutunlara yazar."""
        satir = _UCRET_TABLO_BASLANGIC_SATIR
        for anahtar, deger in BRUT_UCRET_HARITASI.items():
            ws[f"{_UCRET_TABLO_ANAHTAR_SUTUNU}{satir}"] = anahtar
            ws[f"{_UCRET_TABLO_DEGER_SUTUNU}{satir}"] = deger
            satir += 1

        ws.column_dimensions[_UCRET_TABLO_ANAHTAR_SUTUNU].hidden = True
        ws.column_dimensions[_UCRET_TABLO_DEGER_SUTUNU].hidden = True

    @staticmethod
    def _ekle_veri_dogrulama(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Öğrenim ve hizmet grubu türü hücreleri için açılır listeleri ekler."""
        from openpyxl.worksheet.datavalidation import DataValidation
        from src.config.constants import OGRENIM_SEVIYELERI

        # Şablondan M3 için gelen eski doğrulamayı temizleyip yeni A/AG listesini ekleriz.
        if hasattr(ws, "data_validations") and hasattr(
            ws.data_validations, "dataValidation"
        ):
            ws.data_validations.dataValidation = [
                dv
                for dv in ws.data_validations.dataValidation
                if str(dv.sqref) != _HUCRE_HIZMET_GRUBU_TURU
            ]

        ogrenim_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(OGRENIM_SEVIYELERI)}"',
            allow_blank=True,
        )
        ws.add_data_validation(ogrenim_dv)
        for row in range(6, 11):  # 6, 7, 8, 9, 10
            ogrenim_dv.add(f"B{row}")

        hizmet_grubu_dv = DataValidation(
            type="list",
            formula1='"A,AG"',
            allow_blank=True,
        )
        ws.add_data_validation(hizmet_grubu_dv)
        hizmet_grubu_dv.add(_HUCRE_HIZMET_GRUBU_TURU)
