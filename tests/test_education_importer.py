"""education_importer modülü testleri."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.core.education_importer import EducationImporter


def _write_source_xlsx(rows: list[dict[str, str | None]], path: Path) -> None:
    """Kaynak mezuniyet dosyası üretir."""
    dataframe = pd.DataFrame(rows)
    dataframe.to_excel(path, index=False)


def _create_target_workbook(
    path: Path,
    sheet_title: str = "Ali YILMAZ - 10000000146",
) -> None:
    """Hedef tutanak workbook'unu minimal şablonla oluşturur."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet["B4"] = "ÖĞRENİM BİLGİLERİ"
    worksheet["C5"] = "OKUL"
    worksheet["E5"] = "BÖLÜM"
    worksheet["B9"] = "MESLEKİ TECRÜBELER"
    workbook.save(path)
    workbook.close()


class TestEducationImporter:
    """EducationImporter davranış testleri."""

    @pytest.fixture()
    def importer(self) -> EducationImporter:
        return EducationImporter()

    def test_import_education_creates_backup_and_appends_records(
        self,
        importer: EducationImporter,
        tmp_path: Path,
    ):
        """Eşleşen sayfaya mezuniyet kayıtları yazılmalı ve yedek alınmalı."""
        source_path = tmp_path / "mezuniyet.xlsx"
        target_path = tmp_path / "tutanak.xlsx"

        _write_source_xlsx(
            [
                {
                    "TC KIMLIK NO": "10000000146",
                    "AD": "ALİ",
                    "SOYAD": "YILMAZ",
                    "MEZUNIYET TARIHI": "03/01/2022",
                    "UNIVERSITE": "İSTANBUL TEKNİK ÜNİVERSİTESİ",
                    "ENSMYOFAK": "ELEKTRİK-ELEKTRONİK FAKÜLTESİ",
                    "PROGRAM": "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ",
                },
                {
                    "TC KIMLIK NO": "10000000146",
                    "AD": "ALİ",
                    "SOYAD": "YILMAZ",
                    "MEZUNIYET TARIHI": "15/07/2024",
                    "UNIVERSITE": "BOĞAZİÇİ ÜNİVERSİTESİ",
                    "ENSMYOFAK": "FEN BİLİMLERİ ENSTİTÜSÜ",
                    "PROGRAM": "YÜKSEK LİSANS FİZİK",
                },
                {
                    "TC KIMLIK NO": "10000000078",
                    "AD": "VELİ",
                    "SOYAD": "DEMİR",
                    "MEZUNIYET TARIHI": "10/06/2020",
                    "UNIVERSITE": "ORTA DOĞU TEKNİK ÜNİVERSİTESİ",
                    "ENSMYOFAK": "MÜHENDİSLİK FAKÜLTESİ",
                    "PROGRAM": "BİLGİSAYAR MÜHENDİSLİĞİ",
                },
                {
                    "TC KIMLIK NO": "10000050028",
                    "AD": "Mezun Kaydı Bulunamadı !!!",
                    "SOYAD": "",
                    "MEZUNIYET TARIHI": None,
                    "UNIVERSITE": None,
                    "ENSMYOFAK": None,
                    "PROGRAM": None,
                },
            ],
            source_path,
        )
        _create_target_workbook(target_path)

        result = importer.import_education(source_path, target_path)

        assert result.backup_path.exists()
        assert "_eski_" in result.backup_path.name
        assert result.matched_sheet_count == 1
        assert result.updated_sheet_count == 1
        assert result.appended_record_count == 2
        assert result.skipped_record_count == 0
        assert result.unmatched_tckns == ["10000000078"]

        workbook = openpyxl.load_workbook(target_path)
        worksheet = workbook.active
        assert worksheet["B6"].value == "Lisans"
        assert "İSTANBUL TEKNİK ÜNİVERSİTESİ" in worksheet["C6"].value
        assert worksheet["E6"].value == "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ"
        assert worksheet["B7"].value == "Tezli Yüksek Lisans"
        assert "15.07.2024" in worksheet["C7"].value
        workbook.close()

    def test_import_education_preserves_existing_rows_and_skips_duplicates(
        self,
        importer: EducationImporter,
        tmp_path: Path,
    ):
        """Dolu satırlar korunmalı, tekrar eden kayıt atlanmalı."""
        source_path = tmp_path / "mezuniyet.xlsx"
        target_path = tmp_path / "tutanak.xlsx"

        _write_source_xlsx(
            [
                {
                    "TC KIMLIK NO": "10000000146",
                    "AD": "ALİ",
                    "SOYAD": "YILMAZ",
                    "MEZUNIYET TARIHI": "03/01/2022",
                    "UNIVERSITE": "İSTANBUL TEKNİK ÜNİVERSİTESİ",
                    "ENSMYOFAK": "ELEKTRİK-ELEKTRONİK FAKÜLTESİ",
                    "PROGRAM": "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ",
                },
                {
                    "TC KIMLIK NO": "10000000146",
                    "AD": "ALİ",
                    "SOYAD": "YILMAZ",
                    "MEZUNIYET TARIHI": "10/06/2020",
                    "UNIVERSITE": "ORTA DOĞU TEKNİK ÜNİVERSİTESİ",
                    "ENSMYOFAK": "MÜHENDİSLİK FAKÜLTESİ",
                    "PROGRAM": "BİLGİSAYAR MÜHENDİSLİĞİ",
                },
            ],
            source_path,
        )
        _create_target_workbook(target_path)

        workbook = openpyxl.load_workbook(target_path)
        worksheet = workbook.active
        worksheet["B6"] = "Lisans"
        worksheet["C6"] = (
            "İSTANBUL TEKNİK ÜNİVERSİTESİ - "
            "ELEKTRİK-ELEKTRONİK FAKÜLTESİ - "
            "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ - 03.01.2022"
        )
        worksheet["E6"] = "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ"
        workbook.save(target_path)
        workbook.close()

        result = importer.import_education(source_path, target_path)

        assert result.appended_record_count == 1
        assert result.skipped_record_count == 1

        workbook = openpyxl.load_workbook(target_path)
        worksheet = workbook.active
        assert worksheet["C6"].value.startswith("İSTANBUL TEKNİK ÜNİVERSİTESİ")
        assert "ORTA DOĞU TEKNİK ÜNİVERSİTESİ" in worksheet["C7"].value
        workbook.close()

    def test_import_education_raises_for_empty_valid_source(
        self,
        importer: EducationImporter,
        tmp_path: Path,
    ):
        """Tüm satırlar filtrelenirse anlamlı hata yükseltilmeli."""
        source_path = tmp_path / "mezuniyet.xlsx"
        target_path = tmp_path / "tutanak.xlsx"

        _write_source_xlsx(
            [
                {
                    "TC KIMLIK NO": None,
                    "AD": "ALİ",
                    "SOYAD": "YILMAZ",
                    "MEZUNIYET TARIHI": "03/01/2022",
                    "UNIVERSITE": "İSTANBUL TEKNİK ÜNİVERSİTESİ",
                    "ENSMYOFAK": "ELEKTRİK-ELEKTRONİK FAKÜLTESİ",
                    "PROGRAM": "ELEKTRONİK VE HABERLEŞME MÜHENDİSLİĞİ",
                },
                {
                    "TC KIMLIK NO": "10000000146",
                    "AD": "Mezun Kaydı Bulunamadı !!!",
                    "SOYAD": "",
                    "MEZUNIYET TARIHI": None,
                    "UNIVERSITE": None,
                    "ENSMYOFAK": None,
                    "PROGRAM": None,
                },
            ],
            source_path,
        )
        _create_target_workbook(target_path)

        with pytest.raises(
            ValueError,
            match="işlenecek geçerli mezuniyet kaydı",
        ):
            importer.import_education(source_path, target_path)

        assert not list(tmp_path.glob("*_eski_*.xlsx"))

