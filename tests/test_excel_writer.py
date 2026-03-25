"""
excel_writer modülü için entegrasyon testleri.

xlsx bellekte oluşturulur (BytesIO ile), openpyxl ile açılıp
sayfa sayısı, sayfa adları ve hücre değerleri doğrulanır.
"""

from __future__ import annotations

import openpyxl
import pytest

from src.core.excel_reader import Personel
from src.core.excel_writer import (
    _sayfa_adi_olustur,
    olustur_dk_dosyasi_raporlu,
)
from src.core.excel_write_strategy_v1 import ExcelWriteStrategyV1
from src.config.constants import TECRUBE_BASLANGIC_SATIR

_HUCRE_UNVAN = ExcelWriteStrategyV1.HUCRE_UNVAN


# ---------------------------------------------------------------------------
# Fixture'lar
# ---------------------------------------------------------------------------


@pytest.fixture()
def tek_personel() -> list[Personel]:
    return [
        Personel(tckn="10000000146", ad_soyad="Fatma KARACA", birim="Marmara Enstitüsü")
    ]


@pytest.fixture()
def uc_personel() -> list[Personel]:
    return [
        Personel(
            tckn="10000000146", ad_soyad="Fatma KARACA", birim="Marmara Enstitüsü"
        ),
        Personel(tckn="10000000078", ad_soyad="Ali YILMAZ", birim="Gebze Enstitüsü"),
        Personel(tckn="10000050028", ad_soyad="Ayşe DEMİR", birim="Kocaeli Enstitüsü"),
    ]


# ---------------------------------------------------------------------------
# _sayfa_adi_olustur testleri
# ---------------------------------------------------------------------------


class TestSayfaAdiOlustur:
    """_sayfa_adi_olustur fonksiyonu için testler."""

    def test_normal_ad(self):
        p = Personel(tckn="10000000146", ad_soyad="Ali VELİ", birim="Birim")
        sonuc = _sayfa_adi_olustur(p)
        assert "Ali VELİ" in sonuc
        assert "10000000146" in sonuc

    def test_max_31_karakter(self):
        """Sayfa adı 31 karakteri geçmemeli."""
        p = Personel(
            tckn="10000000146",
            ad_soyad="Çok Uzun Bir İsim Soyisim Örneği",
            birim="Birim",
        )
        sonuc = _sayfa_adi_olustur(p)
        assert len(sonuc) <= 31

    def test_gecersiz_karakterler_temizlenir(self):
        """Geçersiz Excel sayfa karakterleri temizlenmeli."""
        p = Personel(tckn="10000000146", ad_soyad="Ali/Veli", birim="Birim")
        sonuc = _sayfa_adi_olustur(p)
        assert "/" not in sonuc


# ---------------------------------------------------------------------------
# olustur_dk_bytes testleri
# ---------------------------------------------------------------------------


class TestOlusturDkDosyasiRaporlu:
    """olustur_dk_dosyasi_raporlu fonksiyonu için testler."""

    def test_tek_personel_tek_sayfa(self, tmp_path, tek_personel):
        """Bir personel için bir sayfa oluşturulmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        gorunen = [s for s in wb.sheetnames if not wb[s].sheet_state == "hidden"]
        assert len(gorunen) == 1
        wb.close()

    def test_uc_personel_uc_sayfa(self, tmp_path, uc_personel):
        """Üç personel için üç sayfa oluşturulmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=uc_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        gorunen = [s for s in wb.sheetnames if wb[s].sheet_state != "hidden"]
        assert len(gorunen) == 3
        wb.close()

    def test_gecerli_xlsx_uretilir(self, tmp_path, tek_personel):
        """Üretilen dosya geçerli xlsx formatında olmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        assert wb is not None
        wb.close()

    def test_ad_soyad_hucrede(self, tmp_path, tek_personel):
        """Ad Soyad değeri B3 hücresine yazılmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws["B3"].value == "Fatma KARACA"
        wb.close()

    def test_tckn_hucrede(self, tmp_path, tek_personel):
        """TCKN değeri C3 hücresine yazılmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws["C3"].value == "10000000146"
        wb.close()

    def test_birim_hucrede(self, tmp_path, tek_personel):
        """Birim değeri D3 hücresine yazılmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws["D3"].value == "Marmara Enstitüsü"
        wb.close()

    def test_hizmet_grubu_turu_varsayilan_ag(self, tmp_path, tek_personel):
        """Hizmet grubu türü seçim hücresine varsayılan AG yazılmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws["M3"].value == "AG"
        wb.close()

    def test_hizmet_grubu_basligi_guncellenir(self, tmp_path, tek_personel):
        """Seçim alanının başlığı hizmet grubu türünü göstermeli."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws["M2"].value == "HİZMET GRUBU TÜRÜ"
        wb.close()

    def test_hizmet_grubu_sutunu_genisler(self, tmp_path, tek_personel):
        """M sütunu başlık metnini sığdıracak kadar geniş olmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        assert ws.column_dimensions["M"].width >= 22
        wb.close()

    def test_tecrube_formulleri_var(self, tmp_path, tek_personel):
        """Tecrübe satırlarında K sütununda formül olmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        k10 = ws.cell(row=TECRUBE_BASLANGIC_SATIR, column=11).value
        assert k10 is not None
        assert str(k10).startswith("=")
        wb.close()

    def test_alanda_formulleri_var(self, tmp_path, tek_personel):
        """Tecrübe satırlarında L sütununda formül olmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        l10 = ws.cell(row=TECRUBE_BASLANGIC_SATIR, column=12).value
        assert l10 is not None
        assert str(l10).startswith("=")
        wb.close()

    def test_sayfa_adi_tckn_icerir(self, tmp_path, tek_personel):
        """Sayfa adı TCKN numarasını içermeli."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        assert any("10000000146" in name for name in wb.sheetnames)
        wb.close()

    def test_bos_liste_sayfa_kleri_bos(self, tmp_path):
        """Boş liste verildiğinde personel sayfası oluşturulmamalı, sadece _bos yer tutucu olmalı."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        # Sayfa adlarında hiçbir geçerli TCKN olmamalı
        personel_sayfalari = [s for s in wb.sheetnames if s != "_bos"]
        assert len(personel_sayfalari) == 0
        wb.close()

    def test_unvan_formul_yazildi(self, tmp_path, tek_personel):
        """Ünvan hücresi formül içermeli."""
        rapor = olustur_dk_dosyasi_raporlu(
            personeller=tek_personel,
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )
        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]
        unvan_hucre = ws[_HUCRE_UNVAN].value
        assert unvan_hucre is not None
        assert str(unvan_hucre).startswith("=")
        wb.close()

    def test_ayni_personel_ikinci_kez_atlanir(self, tmp_path):
        """Aynı sayfa adı tekrar oluşursa atlama raporlanmalı."""
        personeller = [
            Personel(
                tckn="10000000146",
                ad_soyad="Fatma KARACA",
                birim="Marmara Enstitüsü",
            ),
            Personel(
                tckn="10000000146",
                ad_soyad="Fatma KARACA",
                birim="Marmara Enstitüsü",
            ),
        ]

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=personeller,
            cikti_dizini=tmp_path,
            dosya_adi="DK_Tutanaklari.xlsx",
        )

        assert rapor.output_path == tmp_path / "DK_Tutanaklari.xlsx"
        assert rapor.added_sheet_count == 1
        assert rapor.skipped_existing_count == 1
        assert len(rapor.warning_messages) == 1
        assert "hedef dosyada zaten mevcut" in rapor.warning_messages[0]
