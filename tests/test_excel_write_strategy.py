"""
ExcelWriteStrategy, ExcelWriteStrategyV1 ve ExcelWriterFactory testleri.

Strategy Pattern + Factory Method'un doğru çalıştığını doğrular.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from src.core.excel_reader import Personel
from src.core.excel_write_strategy import ExcelWriteStrategy
from src.core.excel_write_strategy_v1 import ExcelWriteStrategyV1
from src.core.excel_writer_factory import ExcelWriterFactory

# ---------------------------------------------------------------------------
# Factory testleri
# ---------------------------------------------------------------------------


class TestExcelWriterFactory:
    """ExcelWriterFactory sınıfı testleri."""

    def test_create_v1(self):
        """v1 versiyonu ExcelWriteStrategyV1 döndürmeli."""
        strategy = ExcelWriterFactory.create("v1")
        assert isinstance(strategy, ExcelWriteStrategyV1)

    def test_create_v1_is_strategy(self):
        """v1 versiyonu ExcelWriteStrategy arayüzünü implemente etmeli."""
        strategy = ExcelWriterFactory.create("v1")
        assert isinstance(strategy, ExcelWriteStrategy)

    def test_create_invalid_version(self):
        """Geçersiz versiyon ValueError fırlatmalı."""
        with pytest.raises(ValueError, match="Desteklenmeyen"):
            ExcelWriterFactory.create("v99")

    def test_create_empty_version(self):
        """Boş versiyon ValueError fırlatmalı."""
        with pytest.raises(ValueError):
            ExcelWriterFactory.create("")


# ---------------------------------------------------------------------------
# V1 Strategy testleri
# ---------------------------------------------------------------------------


class TestExcelWriteStrategyV1:
    """ExcelWriteStrategyV1 sınıfı testleri."""

    @pytest.fixture()
    def strategy(self) -> ExcelWriteStrategyV1:
        return ExcelWriteStrategyV1()

    @pytest.fixture()
    def personel(self) -> Personel:
        return Personel(
            tckn="10000000146",
            ad_soyad="Fatma KARACA",
            birim="Marmara Enstitüsü",
        )

    @pytest.fixture()
    def template_ws(self):
        """Test için basit bir şablon çalışma sayfası oluşturur."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Minimum şablon yapısı – strateji bunu dolduracak
        return ws

    def test_sayfa_doldur_ad_soyad(self, strategy, personel, template_ws):
        """Ad soyad B3 hücresine yazılmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["B3"].value == "Fatma KARACA"

    def test_sayfa_doldur_tckn(self, strategy, personel, template_ws):
        """TCKN C3 hücresine yazılmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["C3"].value == "10000000146"

    def test_sayfa_doldur_birim(self, strategy, personel, template_ws):
        """Birim D3 hücresine yazılmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["D3"].value == "Marmara Enstitüsü"

    def test_sayfa_doldur_hizmet_grubu_basligi(self, strategy, personel, template_ws):
        """Hizmet grubu seçim alanının başlığı güncellenmeli."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["M2"].value == "HİZMET GRUBU TÜRÜ"

    def test_sayfa_doldur_hizmet_grubu_turu_varsayilan_ag(
        self, strategy, personel, template_ws
    ):
        """Hizmet grubu türü seçim hücresi varsayılan olarak AG olmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["M3"].value == "AG"

    def test_sayfa_doldur_hizmet_grubu_hucreleri_komsu_bicimini_alir(
        self, strategy, personel, template_ws
    ):
        """M2 ve M3 görünür kaynak hücrelerin biçimini almalı."""
        template_ws["H2"].fill = PatternFill(fill_type="solid", fgColor="FFCC00")
        template_ws["H3"].fill = PatternFill(fill_type="solid", fgColor="99CCFF")

        strategy.sayfa_doldur(template_ws, personel)

        assert template_ws["M2"]._style == template_ws["H2"]._style
        assert template_ws["M3"]._style == template_ws["H3"]._style

    def test_sayfa_doldur_hizmet_grubu_sutunu_genisler(
        self, strategy, personel, template_ws
    ):
        """M sütunu başlık metnini gösterecek kadar genişlemeli."""
        template_ws.column_dimensions["M"].width = 13

        strategy.sayfa_doldur(template_ws, personel)

        assert template_ws.column_dimensions["M"].width >= 22

    def test_sayfa_doldur_unvan_formul(self, strategy, personel, template_ws):
        """Ünvan hücresi formül içermeli."""
        strategy.sayfa_doldur(template_ws, personel)
        unvan = template_ws["E3"].value
        assert unvan is not None
        assert str(unvan).startswith("=")
        assert 'M3="AG"' in str(unvan)

    def test_sayfa_doldur_kademe_formul(self, strategy, personel, template_ws):
        """Kademe hücresi formül içermeli."""
        strategy.sayfa_doldur(template_ws, personel)
        kademe = template_ws["F3"].value
        assert kademe is not None
        assert str(kademe).startswith("=")

    def test_sayfa_doldur_brut_ucret_formul(self, strategy, personel, template_ws):
        """G3 hücresi F3'e bağlı brüt ücret formülü içermeli."""
        strategy.sayfa_doldur(template_ws, personel)
        brut_ucret = template_ws["G3"].value
        assert brut_ucret is not None
        assert str(brut_ucret).startswith("=")
        assert "VLOOKUP(F3," in str(brut_ucret)

    def test_sayfa_doldur_brut_ucret_tablosu_gizli_sutunlara_yazilir(
        self, strategy, personel, template_ws
    ):
        """Ucret esleme tablosu AA/AB sutunlarina yazilmali ve gizli olmali."""
        strategy.sayfa_doldur(template_ws, personel)
        assert template_ws["AA1"].value == "AG-1/6"
        assert template_ws["AB1"].value == 347991.73
        assert template_ws.column_dimensions["AA"].hidden is True
        assert template_ws.column_dimensions["AB"].hidden is True

    def test_sayfa_doldur_tecrube_formulleri(self, strategy, personel, template_ws):
        """Tecrübe satırlarında formül olmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        from src.config.constants import TECRUBE_BASLANGIC_SATIR

        k_val = template_ws.cell(row=TECRUBE_BASLANGIC_SATIR, column=11).value
        assert k_val is not None
        assert str(k_val).startswith("=")

    def test_sayfa_doldur_360_yil_ay_gun_formulleri(
        self, strategy, personel, template_ws
    ):
        """J29/K29/L29 formülleri L28 bazlı olmalı ve tam sayı formatında olmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        from src.config.constants import TECRUBE_BITIS_SATIR

        satir = TECRUBE_BITIS_SATIR + 2
        yil = template_ws.cell(row=satir, column=10)
        ay = template_ws.cell(row=satir, column=11)
        gun = template_ws.cell(row=satir, column=12)

        for hucre in (yil, ay, gun):
            assert hucre.value is not None
            assert str(hucre.value).startswith("=")
            assert "L28" in str(hucre.value)
            assert hucre.number_format == "0"

    def test_sayfa_doldur_hizmet_grubu_formulu_M3e_bagli(
        self, strategy, personel, template_ws
    ):
        """Hizmet grubu formülü seçim hücresi olarak M3'ü kullanmalı."""
        strategy.sayfa_doldur(template_ws, personel)
        assert 'OR(M3="A",M3="AG")' in str(template_ws["Z2"].value)

    def test_is_abstract_strategy_subclass(self):
        """V1 stratejisi ExcelWriteStrategy alt sınıfı olmalı."""
        assert issubclass(ExcelWriteStrategyV1, ExcelWriteStrategy)
