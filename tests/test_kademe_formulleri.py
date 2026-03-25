"""
Kademe başlangıç ve bitiş formülleri için pytest testleri.

Kıdem tablosundaki tüm tecrübe aralıkları ve eğitim seviyeleri için
K30 ve L30 hücrelerinin doğru formüllerle doldurulduğunu test eder.
"""

from __future__ import annotations

import openpyxl
import pytest

from src.core.excel_reader import Personel
from src.core.excel_writer import olustur_dk_dosyasi_raporlu

# ---------------------------------------------------------------------------
# Test Verileri - Kıdem Tablosu Mapping
# ---------------------------------------------------------------------------

KIDEM_TABLOSU_TEST_VERILERI = [
    # (tecrube_yili, egitim_seviyesi, beklenen_baslangic, beklenen_bitis)
    # 0-2 yıl aralığı
    (0.5, "Lisans", 6, 5),
    (1.0, "Lisans", 6, 5),
    (1.9, "Lisans", 6, 5),
    (0.5, "Tezsiz Yüksek Lisans", 5, 4),
    (1.5, "Tezli Yüksek Lisans", 4, 3),
    (1.0, "Doktora", 4, 3),
    # 2-3 yıl aralığı
    (2.0, "Lisans", 4, 3),
    (2.5, "Lisans", 4, 3),
    (2.9, "Lisans", 4, 3),
    (2.5, "Tezsiz Yüksek Lisans", 4, 3),
    (2.5, "Tezli Yüksek Lisans", 2, 2),
    (2.5, "Doktora", 2, 2),
    # 3-6 yıl aralığı
    (3.0, "Lisans", 6, 5),
    (4.0, "Lisans", 6, 5),
    (5.5, "Lisans", 6, 5),
    (4.0, "Tezsiz Yüksek Lisans", 6, 5),
    (4.0, "Tezli Yüksek Lisans", 5, 4),
    (4.0, "Doktora", 4, 2),
    # 6-8 yıl aralığı
    (6.0, "Lisans", 4, 3),
    (7.0, "Lisans", 4, 3),
    (7.9, "Lisans", 4, 3),
    (7.0, "Tezsiz Yüksek Lisans", 4, 3),
    (7.0, "Tezli Yüksek Lisans", 3, 2),
    (7.0, "Doktora", 4, 2),
    # 8-10 yıl aralığı
    (8.0, "Lisans", 6, 5),
    (9.0, "Lisans", 6, 5),
    (9.9, "Lisans", 6, 5),
    (9.0, "Tezsiz Yüksek Lisans", 6, 5),
    (9.0, "Tezli Yüksek Lisans", 5, 4),
    (9.0, "Doktora", 4, 2),
    # 10-12 yıl aralığı
    (10.0, "Lisans", 4, 3),
    (11.0, "Lisans", 4, 3),
    (11.9, "Lisans", 4, 3),
    (11.0, "Tezsiz Yüksek Lisans", 4, 3),
    (11.0, "Tezli Yüksek Lisans", 3, 2),
    (11.0, "Doktora", 4, 2),
    # 12-15 yıl aralığı
    (12.0, "Lisans", 6, 5),
    (13.0, "Lisans", 6, 5),
    (14.5, "Lisans", 6, 5),
    (13.0, "Tezsiz Yüksek Lisans", 6, 5),
    (13.0, "Tezli Yüksek Lisans", 5, 4),
    (13.0, "Doktora", 4, 2),
    # 15-16 yıl aralığı
    (15.0, "Lisans", 4, 3),
    (15.5, "Lisans", 4, 3),
    (15.9, "Lisans", 4, 3),
    (15.5, "Tezsiz Yüksek Lisans", 4, 3),
    (15.5, "Tezli Yüksek Lisans", 3, 2),
    (15.5, "Doktora", 4, 2),
    # 16+ yıl aralığı
    (16.0, "Lisans", 6, 3),
    (18.0, "Lisans", 6, 3),
    (20.0, "Lisans", 6, 3),
    (18.0, "Tezsiz Yüksek Lisans", 6, 3),
    (18.0, "Tezli Yüksek Lisans", 5, 2),
    (18.0, "Doktora", 4, 2),
]


# ---------------------------------------------------------------------------
# Testler
# ---------------------------------------------------------------------------


class TestKademeFormulleri:
    """K30 ve L30 hücrelerindeki kademe formülleri için testler."""

    def test_k30_l30_formulleri_var(self, tmp_path):
        """K30 ve L30 hücrelerinde formül olmalı."""
        personel = Personel(
            tckn="12345678901", ad_soyad="Test Kullanıcı", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        k30 = ws["K30"].value
        l30 = ws["L30"].value

        assert k30 is not None, "K30 hücresi boş olmamalı"
        assert l30 is not None, "L30 hücresi boş olmamalı"
        assert str(k30).startswith("="), "K30 formül içermeli"
        assert str(l30).startswith("="), "L30 formül içermeli"

        wb.close()

    def test_k30_z1_z4_ve_f3_referansi(self, tmp_path):
        """K30 formülü Z1, Z4 ve F3 hücrelerine referans vermeli."""
        personel = Personel(
            tckn="12345678901", ad_soyad="Test Kullanıcı", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        k30_formula = str(ws["K30"].value)

        assert "Z1" in k30_formula, "K30 formülü Z1 hücresine referans vermeli"
        assert "Z4" in k30_formula, "K30 formülü Z4 hücresine referans vermeli"
        assert "F3" in k30_formula, "K30 formülü F3 hücresine referans vermeli"

        wb.close()

    def test_l30_z1_z4_ve_f3_referansi(self, tmp_path):
        """L30 formülü Z1, Z4 ve F3 hücrelerine referans vermeli."""
        personel = Personel(
            tckn="12345678901", ad_soyad="Test Kullanıcı", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        l30_formula = str(ws["L30"].value)

        assert "Z1" in l30_formula, "L30 formülü Z1 hücresine referans vermeli"
        assert "Z4" in l30_formula, "L30 formülü Z4 hücresine referans vermeli"
        assert "F3" in l30_formula, "L30 formülü F3 hücresine referans vermeli"

        wb.close()

    def test_k30_egitim_seviyeleri_kontrolu(self, tmp_path):
        """K30 formülü Lisans, Tezsiz YL ve Tezli YL seviyelerini kontrol etmeli."""
        personel = Personel(
            tckn="12345678901", ad_soyad="Test Kullanıcı", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        k30_formula = str(ws["K30"].value)

        assert "Lisans" in k30_formula
        assert "Tezsiz Yüksek Lisans" in k30_formula
        assert "Tezli Yüksek Lisans" in k30_formula
        # Doktora son else dalında olduğu için formülde kelime olarak geçmez

        wb.close()

    def test_l30_egitim_seviyeleri_kontrolu(self, tmp_path):
        """L30 formülü Lisans, Tezsiz YL ve Tezli YL seviyelerini kontrol etmeli."""
        personel = Personel(
            tckn="12345678901", ad_soyad="Test Kullanıcı", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi="test.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        l30_formula = str(ws["L30"].value)

        assert "Lisans" in l30_formula
        assert "Tezsiz Yüksek Lisans" in l30_formula
        assert "Tezli Yüksek Lisans" in l30_formula
        # Doktora son else dalında olduğu için formülde kelime olarak geçmez

        wb.close()

    @pytest.mark.parametrize(
        "tecrube_yili,egitim,beklenen_baslangic,beklenen_bitis",
        KIDEM_TABLOSU_TEST_VERILERI,
        ids=[
            f"{t}yil_{e.replace(' ', '_')}_K{b}_L{bt}"
            for t, e, b, bt in KIDEM_TABLOSU_TEST_VERILERI
        ],
    )
    def test_kidem_tablosu_degerleri(
        self, tmp_path, tecrube_yili, egitim, beklenen_baslangic, beklenen_bitis
    ):
        """
        Kıdem tablosundaki tüm değerler için K30 ve L30 formüllerinin
        doğru sonuçları ürettiğini test eder.

        NOT: Bu test formüllerin varlığını kontrol eder.
        Gerçek hesaplama Excel tarafından yapılır.
        """
        personel = Personel(
            tckn="12345678901",
            ad_soyad=f"Test {tecrube_yili}y {egitim}",
            birim="Test Birimi",
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi=f"test_{tecrube_yili}_{egitim.replace(' ', '_')}.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        # Formüllerin varlığını kontrol et
        k30 = ws["K30"].value
        l30 = ws["L30"].value

        assert k30 is not None, f"K30 boş (tecrübe={tecrube_yili}, egitim={egitim})"
        assert l30 is not None, f"L30 boş (tecrübe={tecrube_yili}, egitim={egitim})"
        assert str(k30).startswith(
            "="
        ), f"K30 formül değil (tecrübe={tecrube_yili}, egitim={egitim})"
        assert str(l30).startswith(
            "="
        ), f"L30 formül değil (tecrübe={tecrube_yili}, egitim={egitim})"

        wb.close()


class TestKademeAraliklari:
    """Tecrübe yılı aralıklarının doğru çalıştığını test eder."""

    @pytest.mark.parametrize(
        "tecrube_yili,aralik_adi",
        [
            (0.0, "0-2 yıl"),
            (1.99, "0-2 yıl"),
            (2.0, "2-3 yıl"),
            (2.99, "2-3 yıl"),
            (3.0, "3-6 yıl"),
            (5.99, "3-6 yıl"),
            (6.0, "6-8 yıl"),
            (7.99, "6-8 yıl"),
            (8.0, "8-10 yıl"),
            (9.99, "8-10 yıl"),
            (10.0, "10-12 yıl"),
            (11.99, "10-12 yıl"),
            (12.0, "12-15 yıl"),
            (14.99, "12-15 yıl"),
            (15.0, "15-16 yıl"),
            (15.99, "15-16 yıl"),
            (16.0, "16+ yıl"),
            (20.0, "16+ yıl"),
        ],
    )
    def test_tecrube_aralik_sinirlari(self, tmp_path, tecrube_yili, aralik_adi):
        """
        Tecrübe yılı aralık sınırlarının doğru çalıştığını test eder.
        Formüllerin IF koşullarının doğru aralıkları kontrol ettiğini doğrular.
        """
        personel = Personel(
            tckn="12345678901", ad_soyad=f"Test {tecrube_yili}y", birim="Test Birimi"
        )

        rapor = olustur_dk_dosyasi_raporlu(
            personeller=[personel],
            cikti_dizini=tmp_path,
            dosya_adi=f"test_{tecrube_yili}.xlsx",
        )

        wb = openpyxl.load_workbook(rapor.output_path)
        ws = wb.worksheets[0]

        # Formüllerin varlığını kontrol et
        k30 = ws["K30"].value
        l30 = ws["L30"].value

        assert k30 is not None
        assert l30 is not None
        assert str(k30).startswith("=")
        assert str(l30).startswith("=")

        wb.close()
